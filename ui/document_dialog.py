# =============================================================================
#  ui/document_dialog.py  –  Creazione / modifica documento
# =============================================================================
import shutil
import subprocess
from pathlib import Path
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QComboBox, QTextEdit, QGroupBox, QFormLayout,
    QTabWidget, QWidget, QTableWidget, QTableWidgetItem,
    QMessageBox, QFileDialog, QHeaderView, QRadioButton, QButtonGroup
)
from PyQt6.QtCore import Qt

from typing import Optional

from config import SW_EXTENSIONS
from ui.session import session
from ui.styles import STATE_BADGE_STYLE, TYPE_ICON


class DocumentDialog(QDialog):
    """Dialogo per creare o visualizzare/modificare un documento."""

    def __init__(self, document_id: Optional[int] = None, parent=None):
        super().__init__(parent)
        self.document_id = document_id
        self.is_new      = document_id is None
        self.setWindowTitle("Nuovo Documento" if self.is_new else "Dettaglio Documento")
        self.setMinimumSize(680, 560)
        self._build_ui()
        if not self.is_new:
            self._load_document()

    # ------------------------------------------------------------------
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(16, 16, 16, 16)

        # Tab
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        self.tabs.addTab(self._build_general_tab(), "Generale")
        self.tabs.addTab(self._build_properties_tab(), "Proprietà SW")
        if not self.is_new:
            self._bom_tab_idx = self.tabs.count()
            self.tabs.addTab(self._build_bom_tab(), "Struttura (BOM)")
            self.tabs.addTab(self._build_history_tab(), "Storico")
            self.tabs.addTab(self._build_revisions_tab(), "Revisioni")

        # Bottoni
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        if self.is_new:
            btn_save = QPushButton("Crea Documento")
            btn_save.setObjectName("btn_primary")
            btn_save.clicked.connect(self._create)
            self.btn_create = btn_save
        else:
            btn_save = QPushButton("Salva Modifiche")
            btn_save.setObjectName("btn_primary")
            btn_save.clicked.connect(self._save)
        btn_close = QPushButton("Chiudi")
        btn_close.clicked.connect(self.accept)
        btn_row.addWidget(btn_close)
        btn_row.addWidget(btn_save)
        layout.addLayout(btn_row)

    # ---- Tab GENERALE ----
    def _build_general_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setSpacing(10)

        grp = QGroupBox("Informazioni documento")
        form = QFormLayout(grp)

        if self.is_new:
            # ---- Selezione gerarchica (solo creazione) ----
            self.cmb_machine = QComboBox()
            self.cmb_machine.setMinimumWidth(180)
            self.cmb_machine.currentIndexChanged.connect(self._on_machine_changed)
            form.addRow("Macchina:", self.cmb_machine)

            self.cmb_group = QComboBox()
            self.cmb_group.setMinimumWidth(180)
            self.cmb_group.currentIndexChanged.connect(self._update_code_preview)
            form.addRow("Gruppo:", self.cmb_group)

            # Radio buttons livello
            self._level_btn_group = QButtonGroup(self)
            level_widget = QWidget()
            level_layout = QVBoxLayout(level_widget)
            level_layout.setContentsMargins(0, 0, 0, 0)
            level_layout.setSpacing(4)
            for _label, _data in [
                ("LIV0  Macchina (ASM)", 0),
                ("LIV1  Gruppo (ASM)", 1),
                ("LIV2/1  Sottogruppo (ASM)", 2),
                ("LIV2/2  Parte (PRT)", 3),
            ]:
                rb = QRadioButton(_label)
                self._level_btn_group.addButton(rb, _data)
                level_layout.addWidget(rb)
            self._level_btn_group.button(0).setChecked(True)
            self._level_btn_group.idClicked.connect(self._on_level_changed)
            form.addRow("Livello:", level_widget)

            # Anteprima codice auto
            self.lbl_code_preview = QLabel("—")
            self.lbl_code_preview.setStyleSheet(
                "color:#89b4fa;font-weight:bold;font-size:14px;"
            )
            form.addRow("Codice generato:", self.lbl_code_preview)

            # Radio buttons modalità creazione
            self._creation_mode_group = QButtonGroup(self)
            mode_widget = QWidget()
            mode_layout = QVBoxLayout(mode_widget)
            mode_layout.setContentsMargins(0, 0, 0, 0)
            mode_layout.setSpacing(4)
            for _mlabel, _mid in [
                ("Crea solo codice", 0),
                ("Crea documento (ASM/PRT)", 1),
                ("Crea documenti (ASM/PRT e DRW)", 2),
            ]:
                rb = QRadioButton(_mlabel)
                self._creation_mode_group.addButton(rb, _mid)
                mode_layout.addWidget(rb)
            self._creation_mode_group.button(1).setChecked(True)
            form.addRow("Crea:", mode_widget)

            # Tipo nascosto (usato internamente)
            self.cmb_type = QComboBox()  # stub per compat.
        else:
            # Visualizzazione documento esistente
            self.lbl_code_preview = QLabel("")
            self.lbl_code_preview.setStyleSheet("font-weight:bold;color:#89b4fa;")
            form.addRow("Codice:", self.lbl_code_preview)
            self.lbl_type_display = QLabel()

        # Revisione
        self.txt_rev = QLineEdit("00")
        self.txt_rev.setMaximumWidth(60)
        self.txt_rev.setPlaceholderText("00")
        if not self.is_new:
            form.addRow("Tipo:", self.lbl_type_display)
        form.addRow("Revisione:", self.txt_rev)

        # Titolo
        self.txt_title = QLineEdit()
        self.txt_title.setPlaceholderText("Descrizione breve del documento")
        form.addRow("Titolo:", self.txt_title)

        # Descrizione
        self.txt_desc = QTextEdit()
        self.txt_desc.setMaximumHeight(80)
        self.txt_desc.setPlaceholderText("Note aggiuntive...")
        form.addRow("Descrizione:", self.txt_desc)

        layout.addWidget(grp)

        # Carica macchine se nuovo documento
        if self.is_new:
            self._load_machines_combo()

        # File
        if not self.is_new:
            grp_file = QGroupBox("File")
            file_form = QFormLayout(grp_file)
            self.lbl_file = QLabel("—")
            file_form.addRow("File in archivio:", self.lbl_file)

            file_row = QHBoxLayout()
            btn_create_sw = QPushButton("\u2728 Crea in SW")
            btn_create_sw.setObjectName("btn_primary")
            btn_create_sw.setToolTip(
                "Crea file da template nella workspace e apre in SolidWorks"
            )
            btn_create_sw.clicked.connect(self._create_from_template)

            btn_copy_file = QPushButton("\U0001f4cb Crea da file")
            btn_copy_file.setToolTip(
                "Copia un file esistente nella workspace rinominandolo col codice PDM"
            )
            btn_copy_file.clicked.connect(self._copy_from_file)

            btn_import = QPushButton("\U0001f4e5 Importa in PDM")
            btn_import.setToolTip(
                "Cerca il file nella workspace e lo archivia nel PDM"
            )
            btn_import.clicked.connect(self._import_from_pdm)

            btn_export = QPushButton("\U0001f4e4 Esporta in WS")
            btn_export.setToolTip(
                "Copia il file dall'archivio nella workspace locale"
            )
            btn_export.clicked.connect(self._export_file)

            btn_open = QPushButton("\U0001f527 Apri in SW")
            btn_open.setToolTip(
                "Copia dall'archivio in workspace (se necessario) e apre in SolidWorks"
            )
            btn_open.clicked.connect(self._open_sw)

            file_row.addWidget(btn_create_sw)
            file_row.addWidget(btn_copy_file)
            file_row.addWidget(btn_import)
            file_row.addWidget(btn_export)
            file_row.addWidget(btn_open)
            file_row.addStretch()
            file_form.addRow(file_row)
            layout.addWidget(grp_file)

            # Stato
            grp_state = QGroupBox("Stato workflow")
            state_form = QFormLayout(grp_state)
            self.lbl_state = QLabel("—")
            self.lbl_locked = QLabel("—")
            state_form.addRow("Stato:", self.lbl_state)
            state_form.addRow("Checkout:", self.lbl_locked)

            # Documento companion (DRW ↔ PRT/ASM)
            self.grp_companion = QGroupBox("Documento collegato")
            companion_layout = QHBoxLayout(self.grp_companion)
            self.lbl_companion = QLabel("—")
            self.btn_companion = QPushButton("")
            self.btn_companion.setObjectName("btn_primary")
            self.btn_companion.setVisible(False)
            companion_layout.addWidget(self.lbl_companion)
            companion_layout.addWidget(self.btn_companion)
            companion_layout.addStretch()
            state_form.addRow("Associato:", self.grp_companion)

            layout.addWidget(grp_state)

        layout.addStretch()
        return w

    # ---- Tab PROPRIETÀ ----
    def _build_properties_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)

        btn_row = QHBoxLayout()
        self.btn_add_prop = QPushButton("+ Aggiungi")
        self.btn_add_prop.clicked.connect(self._add_property_row)
        self.btn_del_prop = QPushButton("- Rimuovi")
        self.btn_del_prop.clicked.connect(self._remove_property_row)
        self.btn_import_sw = QPushButton("Importa da SW")
        self.btn_import_sw.setToolTip(
            "Legge le proprietà custom dal file SolidWorks archiviato\n"
            "(non richiede SolidWorks aperto)"
        )
        self.btn_import_sw.clicked.connect(self._import_props_from_sw)
        self.btn_export_xl = QPushButton("Esporta Excel")
        self.btn_export_xl.clicked.connect(self._export_props_excel)
        self.btn_import_xl = QPushButton("Importa Excel")
        self.btn_import_xl.clicked.connect(self._import_props_excel)
        self.btn_save_props = QPushButton("Salva Proprietà")
        self.btn_save_props.setObjectName("btn_primary")
        self.btn_save_props.clicked.connect(self._save_properties)
        for b in [self.btn_add_prop, self.btn_del_prop, self.btn_import_sw,
                  self.btn_export_xl, self.btn_import_xl, self.btn_save_props]:
            btn_row.addWidget(b)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self.tbl_props = QTableWidget(0, 2)
        self.tbl_props.setHorizontalHeaderLabels(["Proprietà", "Valore"])
        self.tbl_props.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        layout.addWidget(self.tbl_props)
        return w

    # ---- Tab BOM ----
    def _build_bom_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)

        btn_row = QHBoxLayout()
        self.btn_add_comp = QPushButton("+ Aggiungi componente")
        self.btn_add_comp.clicked.connect(self._add_component)
        self.btn_del_comp = QPushButton("- Rimuovi")
        self.btn_del_comp.clicked.connect(self._del_component)
        self.btn_import_asm = QPushButton("Importa struttura da SW")
        self.btn_import_asm.setToolTip(
            "Importa la BOM dall'assieme (legge dal file archiviato,\n"
            "non richiede SolidWorks aperto)"
        )
        self.btn_import_asm.clicked.connect(self._import_asm_from_sw)
        btn_export_bom = QPushButton("📊 Esporta BOM Excel")
        btn_export_bom.clicked.connect(self._export_bom_excel)
        for b in [self.btn_add_comp, self.btn_del_comp, self.btn_import_asm, btn_export_bom]:
            btn_row.addWidget(b)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self.tbl_bom = QTableWidget(0, 5)
        self.tbl_bom.setHorizontalHeaderLabels(
            ["Codice", "Rev.", "Tipo", "Titolo", "Qtà"]
        )
        self.tbl_bom.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeMode.Stretch
        )
        layout.addWidget(self.tbl_bom)

        # ── Sezione articoli commerciali nella BOM ─────────────────────
        from PyQt6.QtWidgets import QGroupBox
        grp_comm = QGroupBox("Articoli Commerciali / Normalizzati")
        comm_lay = QVBoxLayout(grp_comm)

        comm_btn_row = QHBoxLayout()
        self.btn_add_commercial = QPushButton("+ Aggiungi articolo commerciale")
        self.btn_add_commercial.clicked.connect(self._add_commercial_component)
        self.btn_del_commercial = QPushButton("- Rimuovi")
        self.btn_del_commercial.clicked.connect(self._del_commercial_component)
        comm_btn_row.addWidget(self.btn_add_commercial)
        comm_btn_row.addWidget(self.btn_del_commercial)
        comm_btn_row.addStretch()
        comm_lay.addLayout(comm_btn_row)

        self.tbl_bom_commercial = QTableWidget(0, 5)
        self.tbl_bom_commercial.setHorizontalHeaderLabels(
            ["Codice", "Tipo", "Descrizione", "Fornitore pref.", "Qtà"]
        )
        self.tbl_bom_commercial.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch
        )
        self.tbl_bom_commercial.setMaximumHeight(160)
        self.tbl_bom_commercial.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )
        comm_lay.addWidget(self.tbl_bom_commercial)
        layout.addWidget(grp_comm)

        return w

    # ---- Tab STORICO ----
    def _build_history_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        self.tbl_history = QTableWidget(0, 5)
        self.tbl_history.setHorizontalHeaderLabels(
            ["Data/Ora", "Azione", "Da", "A", "Utente"]
        )
        self.tbl_history.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        layout.addWidget(self.tbl_history)
        return w

    # ---- Tab REVISIONI ----
    def _build_revisions_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)

        # Barra pulsanti apertura file revisione
        btn_bar = QHBoxLayout()
        self.btn_rev_edrawings = QPushButton("👁️ Apri in eDrawings")
        self.btn_rev_sw        = QPushButton("🔧 Apri in SW")
        self.btn_rev_pdf       = QPushButton("📄 Apri PDF")
        for b in (self.btn_rev_edrawings, self.btn_rev_sw, self.btn_rev_pdf):
            b.setEnabled(False)
            btn_bar.addWidget(b)
        btn_bar.addStretch()
        layout.addLayout(btn_bar)

        self.tbl_revisions = QTableWidget(0, 4)
        self.tbl_revisions.setHorizontalHeaderLabels(
            ["Revisione", "Stato", "Data rilascio", "Data obsolescenza"]
        )
        hdr = self.tbl_revisions.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.tbl_revisions.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl_revisions.setAlternatingRowColors(True)
        self.tbl_revisions.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        layout.addWidget(self.tbl_revisions)

        self.tbl_revisions.itemSelectionChanged.connect(self._on_revision_selection_changed)
        self.btn_rev_edrawings.clicked.connect(lambda: self._open_revision_file("edrawings"))
        self.btn_rev_sw.clicked.connect(lambda: self._open_revision_file("sw"))
        self.btn_rev_pdf.clicked.connect(lambda: self._open_revision_file("pdf"))
        return w

    def _refresh_revisions(self):
        if not hasattr(self, "tbl_revisions") or not self.document_id:
            return
        doc = session.files.get_document(self.document_id)
        if not doc:
            return
        self.tbl_revisions.setRowCount(0)
        all_revs = session.db.fetchall(
            "SELECT * FROM documents WHERE code=? AND doc_type=? ORDER BY revision DESC",
            (doc["code"], doc["doc_type"]),
        )
        for rd in all_revs:
            hist = session.workflow.get_history(rd["id"])
            release_at = next(
                (h["changed_at"] for h in sorted(hist, key=lambda h: h["changed_at"])
                 if h.get("to_state") == "Rilasciato"), "—"
            )
            obsolete_at = next(
                (h["changed_at"] for h in sorted(hist, key=lambda h: h["changed_at"])
                 if h.get("to_state") == "Obsoleto"), "—"
            )
            i = self.tbl_revisions.rowCount()
            self.tbl_revisions.insertRow(i)
            item_rev = QTableWidgetItem(rd["revision"])
            item_rev.setData(Qt.ItemDataRole.UserRole, rd["id"])
            self.tbl_revisions.setItem(i, 0, item_rev)
            self.tbl_revisions.setItem(i, 1, QTableWidgetItem(rd["state"]))
            self.tbl_revisions.setItem(i, 2, QTableWidgetItem(str(release_at)[:19]))
            self.tbl_revisions.setItem(i, 3, QTableWidgetItem(
                str(obsolete_at)[:19] if obsolete_at != "—" else "—"
            ))

    # ------------------------------------------------------------------
    #  Revisioni — apertura file
    # ------------------------------------------------------------------
    def _get_selected_revision_doc(self) -> "dict | None":
        if not self.tbl_revisions.selectedItems():
            return None
        row = self.tbl_revisions.currentRow()
        item = self.tbl_revisions.item(row, 0)
        if not item:
            return None
        doc_id = item.data(Qt.ItemDataRole.UserRole)
        return session.files.get_document(doc_id) if doc_id and session.files else None

    def _get_pdf_path_for_doc(self, doc: dict) -> "Path | None":
        pdf_path = doc.get("pdf_path") or ""
        if not pdf_path:
            return None
        if session.sp:
            p = session.sp.root / pdf_path
            if p.exists():
                return p
        p_abs = Path(pdf_path)
        return p_abs if p_abs.exists() else None

    def _on_revision_selection_changed(self):
        doc = self._get_selected_revision_doc()
        has_archive = bool(
            doc and doc.get("archive_path") and session.sp
            and (session.sp.root / doc["archive_path"]).exists()
        )
        has_pdf = bool(doc and self._get_pdf_path_for_doc(doc))
        self.btn_rev_edrawings.setEnabled(has_archive)
        self.btn_rev_sw.setEnabled(has_archive)
        self.btn_rev_pdf.setEnabled(has_pdf)

    def _open_revision_file(self, action: str):
        """Copia il file della revisione nella cartella OBSOLETI e lo apre."""
        doc = self._get_selected_revision_doc()
        if not doc:
            return

        from config import load_local_config
        ws = load_local_config().get("sw_workspace", "")
        if not ws:
            QMessageBox.warning(
                self, "Workspace non configurata",
                "Configurare la workspace locale in Strumenti → Impostazioni\n"
                "prima di aprire i file."
            )
            return

        obsoleti_dir = Path(ws) / "OBSOLETI"
        try:
            obsoleti_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            QMessageBox.critical(self, "Errore cartella OBSOLETI", str(e))
            return

        if action in ("edrawings", "sw"):
            src = session.sp.root / doc["archive_path"]
            if not src.exists():
                QMessageBox.warning(self, "File non trovato",
                                    f"Il file non è stato trovato nell'archivio:\n{src}")
                return
            dest = obsoleti_dir / src.name
            try:
                shutil.copy2(src, dest)
            except Exception as e:
                QMessageBox.critical(self, "Errore copia file", str(e))
                return

            if action == "edrawings":
                from ui.sw_config_dialog import SWConfigDialog
                exe = SWConfigDialog.get_edrawings_exe()
                if exe:
                    subprocess.Popen([str(exe), str(dest)])
                else:
                    QMessageBox.warning(
                        self, "eDrawings non configurato",
                        "Eseguibile eDrawings non configurato.\n\n"
                        "Aprire Strumenti → Configurazione SolidWorks → tab SolidWorks\n"
                        "e impostare il percorso di eDrawings."
                    )
            else:
                from config import SW_EXTENSIONS
                doc_type_map = {"Parte": 1, "Assieme": 2, "Disegno": 3}
                type_id = doc_type_map.get(doc.get("doc_type", "Parte"), 1)
                try:
                    import win32com.client
                    try:
                        sw = win32com.client.GetActiveObject("SldWorks.Application")
                    except Exception:
                        sw = win32com.client.Dispatch("SldWorks.Application")
                    sw.Visible = True
                    sw.OpenDoc(str(dest).replace("/", "\\"), type_id)
                except Exception as e:
                    try:
                        from ui.sw_config_dialog import SWConfigDialog
                        sw_exe = SWConfigDialog.get_solidworks_exe()
                        if sw_exe and Path(str(sw_exe)).exists():
                            subprocess.Popen([str(sw_exe), str(dest)])
                        else:
                            subprocess.Popen(["cmd", "/c", "start", "", str(dest)])
                    except Exception:
                        QMessageBox.critical(self, "Errore apertura SolidWorks", str(e))

        elif action == "pdf":
            src = self._get_pdf_path_for_doc(doc)
            if not src:
                QMessageBox.warning(self, "PDF non trovato", "Il file PDF non è disponibile.")
                return
            dest = obsoleti_dir / src.name
            try:
                shutil.copy2(src, dest)
            except Exception as e:
                QMessageBox.critical(self, "Errore copia PDF", str(e))
                return
            import os
            os.startfile(str(dest))

    # ------------------------------------------------------------------
    # Logica
    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    # Gestori selezione gerarchica (solo creazione)
    # ------------------------------------------------------------------
    def _load_machines_combo(self):
        machines = session.coding.get_machines(only_active=True)
        self.cmb_machine.blockSignals(True)
        self.cmb_machine.clear()
        for m in machines:
            self.cmb_machine.addItem(
                f"{m['code']} – {m['description'] or ''}", m["id"]
            )
        self.cmb_machine.blockSignals(False)
        self._on_machine_changed()

    def _on_machine_changed(self):
        machine_id = self.cmb_machine.currentData()
        # Aggiorna combo gruppo
        groups = session.coding.get_groups(machine_id) if machine_id else []
        self.cmb_group.blockSignals(True)
        self.cmb_group.clear()
        for g in groups:
            self.cmb_group.addItem(
                f"{g['code']} – {g['description'] or ''}", g["id"]
            )
        self.cmb_group.blockSignals(False)
        self._update_code_preview()

    def _on_level_changed(self):
        level = self._level_btn_group.checkedId()
        # Nascondi gruppo solo per LIV0 (macchina)
        self.cmb_group.setVisible(level != 0)
        self._update_code_preview()

    def _update_code_preview(self):
        if not self.is_new:
            return
        machine_id = self.cmb_machine.currentData()
        group_id   = self.cmb_group.currentData()
        level_data = self._level_btn_group.checkedId()
        level_map  = {0: (0, "ASM"), 1: (1, "ASM"), 2: (2, "PRT"), 3: (2, "ASM")}
        level, subtype = level_map.get(level_data, (2, "PRT"))
        if not machine_id:
            self.lbl_code_preview.setText("—")
            return
        try:
            preview = session.coding.preview_code(
                level, subtype, machine_id,
                group_id if level > 0 else None
            )
            self.lbl_code_preview.setText(preview)
        except Exception:
            self.lbl_code_preview.setText("—")

    def _gen_code(self):
        """Stub mantenuto per compatibilità."""
        pass

    # ------------------------------------------------------------------
    def _create(self):
        revision   = self.txt_rev.text().strip() or "00"
        title      = self.txt_title.text().strip()
        desc       = self.txt_desc.toPlainText().strip()
        machine_id = self.cmb_machine.currentData()
        group_id   = self.cmb_group.currentData()
        level_data = self._level_btn_group.checkedId()

        if not title:
            QMessageBox.warning(self, "Errore", "Il titolo è obbligatorio")
            return
        if not machine_id:
            QMessageBox.warning(self, "Errore", "Seleziona una macchina")
            return
        if level_data in (1, 2, 3) and not group_id:
            QMessageBox.warning(self, "Errore", "Seleziona un gruppo")
            return

        parent_doc_id = None

        # Mappa livello → (func, doc_type, doc_level)
        try:
            if level_data == 0:
                code      = session.coding.next_code_liv0(machine_id)
                doc_type  = "Assieme"
                doc_level = 0
                group_id  = None
            elif level_data == 1:
                code      = session.coding.next_code_liv1(machine_id, group_id)
                doc_type  = "Assieme"
                doc_level = 1
            elif level_data == 2:
                code      = session.coding.next_code_liv2_part(machine_id, group_id)
                doc_type  = "Assieme"
                doc_level = 2
            elif level_data == 3:
                code      = session.coding.next_code_liv2_subgroup(machine_id, group_id)
                doc_type  = "Parte"
                doc_level = 2
        except Exception as e:
            QMessageBox.critical(self, "Errore generazione codice", str(e))
            return

        if not session.coding.is_code_available(code, revision, doc_type):
            QMessageBox.warning(
                self, "Codice duplicato",
                f"Il codice {code} rev.{revision} [{doc_type}] è già presente"
            )
            return

        doc_id = session.files.create_document(
            code, revision, doc_type, title, desc,
            machine_id=machine_id, group_id=group_id,
            doc_level=doc_level, parent_doc_id=parent_doc_id
        )
        self.document_id = doc_id
        self.is_new = False

        creation_mode = self._creation_mode_group.checkedId()  # 0=solo codice, 1=doc, 2=doc+drw

        if creation_mode == 0:
            # Solo codice: nessun file SW
            QMessageBox.information(
                self, "Documento Creato",
                f"Codice registrato:\n{code}  rev.{revision}  [{doc_type}]"
            )
            self._reset_for_new()
            return

        # Modalità 1 o 2: crea file SW da template
        from config import load_local_config
        from core.file_manager import EXT_FOR_TYPE, _sw_open_and_saveas
        cfg     = load_local_config()
        ws_root = cfg.get("sw_workspace", "")
        key_map = {"Parte": "sw_template_prt", "Assieme": "sw_template_asm", "Disegno": "sw_template_drw"}
        tpl_path = cfg.get(key_map.get(doc_type, ""), "")
        sw_ok = False

        if ws_root and tpl_path and Path(ws_root).is_dir() and Path(tpl_path).exists():
            if creation_mode == 1:
                # Chiedi conferma
                r = QMessageBox.question(
                    self, "Crea file in SolidWorks?",
                    f"Documento creato: <b>{code}  rev.{revision}</b><br><br>"
                    "Aprire il template in SolidWorks e creare il file?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes,
                )
                do_sw = (r == QMessageBox.StandardButton.Yes)
            else:
                do_sw = True  # modalità 2: crea sempre

            if do_sw:
                try:
                    ext = EXT_FOR_TYPE.get(doc_type, ".SLDPRT")
                    ws_file = Path(ws_root) / f"{code}{ext}"
                    _sw_open_and_saveas(Path(tpl_path), ws_file, doc_type, is_template=True)
                    session.checkout.checkout_new_from_workspace(doc_id, ws_file)
                    # Esporta le proprietà PDM nel file appena creato
                    try:
                        session.properties.sync_pdm_to_sw(doc_id, ws_file)
                    except Exception:
                        pass  # non bloccante
                    sw_ok = True
                except Exception as fe:
                    QMessageBox.warning(self, "File non creato",
                                        f"Documento registrato ma file non creato in SW:\n{fe}")

        # Modalità 2: crea anche DRW
        drw_msg = ""
        if creation_mode == 2 and sw_ok:
            drw_tpl_path = cfg.get("sw_template_drw", "")
            if ws_root and drw_tpl_path and Path(drw_tpl_path).exists():
                try:
                    drw_id = session.files.get_or_create_drw_document(doc_id)
                    drw_ext = EXT_FOR_TYPE.get("Disegno", ".SLDDRW")
                    drw_ws_file = Path(ws_root) / f"{code}{drw_ext}"
                    _sw_open_and_saveas(Path(drw_tpl_path), drw_ws_file, "Disegno", is_template=True)
                    session.checkout.checkout_new_from_workspace(drw_id, drw_ws_file)
                    # Esporta le proprietà PDM (del padre PRT/ASM) nel file DRW
                    try:
                        session.properties.sync_pdm_to_sw(drw_id, drw_ws_file)
                    except Exception:
                        pass  # non bloccante
                    drw_msg = f"\n\nDRW creato e in checkout:\n{drw_ws_file}"
                except Exception as fe:
                    QMessageBox.warning(self, "DRW non creato",
                                        f"Documento DRW registrato ma file non creato in SW:\n{fe}")

        if sw_ok:
            sw_ext = EXT_FOR_TYPE.get(doc_type, ".SLDPRT")
            ws_file_path = Path(ws_root) / f"{code}{sw_ext}"
            QMessageBox.information(
                self, "Documento Creato",
                f"Documento creato e in checkout:\n{code}  rev.{revision}  [{doc_type}]"
                f"\n\nFile aperto in SolidWorks dalla workspace:\n{ws_file_path}"
                + drw_msg
            )
        else:
            QMessageBox.information(
                self, "Documento Creato",
                f"Documento creato:\n{code}  rev.{revision}  [{doc_type}]" + drw_msg
            )
        self._reset_for_new()

    def _reset_for_new(self):
        """Azzera titolo e descrizione e ripristina lo stato 'nuovo documento'."""
        self.document_id = None
        self.is_new = True
        self.txt_title.clear()
        self.txt_desc.clear()
        self.txt_title.setFocus()
        self._update_code_preview()

    def _save(self):
        title = self.txt_title.text().strip()
        desc  = self.txt_desc.toPlainText().strip()
        if not title:
            QMessageBox.warning(self, "Errore", "Il titolo è obbligatorio")
            return
        session.files.update_document(self.document_id, title, desc)
        QMessageBox.information(self, "OK", "Documento aggiornato")

    # ------------------------------------------------------------------
    def _load_document(self):
        doc = session.files.get_document(self.document_id)
        if not doc:
            return
        doc_type = doc["doc_type"]
        self.lbl_type_display.setText(f"{TYPE_ICON.get(doc_type, '')}  {doc_type}")
        self.lbl_code_preview.setText(doc["code"])
        self.txt_rev.setText(doc["revision"])
        self.txt_rev.setReadOnly(True)
        self.txt_title.setText(doc["title"])
        self.txt_desc.setPlainText(doc["description"] or "")

        state    = doc["state"]
        badge    = STATE_BADGE_STYLE.get(state, "")
        self.lbl_state.setText(f"<span style='{badge}'>{state}</span>")
        self.lbl_state.setTextFormat(Qt.TextFormat.RichText)

        if doc["is_locked"]:
            locked_name = doc.get("locked_by_name", "sconosciuto")
            self.lbl_locked.setText(
                f"⚠️  In checkout: {locked_name} ({doc.get('locked_ws','')})"
            )
            self.lbl_locked.setStyleSheet("color:#fab387;")
        else:
            self.lbl_locked.setText("✅  Disponibile")
            self.lbl_locked.setStyleSheet("color:#a6e3a1;")

        if doc.get("file_name"):
            self.lbl_file.setText(
                doc["file_name"] + (f"  [{doc['archive_path']}]"
                                    if doc.get("archive_path") else "  (non archiviato)")
            )

        # Proprietà
        props = session.properties.get_properties(self.document_id)
        self.tbl_props.setRowCount(0)
        for name, value in props.items():
            row = self.tbl_props.rowCount()
            self.tbl_props.insertRow(row)
            self.tbl_props.setItem(row, 0, QTableWidgetItem(name))
            self.tbl_props.setItem(row, 1, QTableWidgetItem(value))

        # BOM: visibile solo per Assieme
        if hasattr(self, "_bom_tab_idx"):
            self.tabs.setTabVisible(self._bom_tab_idx, doc_type == "Assieme")

        # BOM
        self._refresh_bom()
        # Storico
        self._refresh_history()
        # Revisioni
        self._refresh_revisions()
        # Companion DRW ↔ PRT/ASM
        self._refresh_companion(doc)

        # Disabilita comandi SW per stati definitivi (tranne Amministratore)
        is_admin = session.can("admin")
        sw_locked = state in ("Rilasciato", "Obsoleto") and not is_admin
        for btn in (self.btn_add_prop, self.btn_del_prop, self.btn_import_sw,
                    self.btn_import_xl, self.btn_save_props):
            btn.setEnabled(not sw_locked)
        if hasattr(self, "btn_add_comp"):
            for btn in (self.btn_add_comp, self.btn_del_comp, self.btn_import_asm):
                btn.setEnabled(not sw_locked)

    # ------------------------------------------------------------------
    def _refresh_companion(self, doc: dict):
        """Aggiorna la sezione documento collegato (DRW ↔ PRT/ASM)."""
        if not hasattr(self, "lbl_companion"):
            return
        doc_type = doc["doc_type"]
        code     = doc["code"]
        revision = doc["revision"]

        if doc_type in ("Parte", "Assieme"):
            # Cerca DRW con stessa revisione
            drw = session.db.fetchone(
                "SELECT * FROM documents WHERE code=? AND doc_type='Disegno' AND revision=?",
                (code, revision),
            )
            if drw:
                state_icon = {"In Lavorazione": "🔵", "In Revisione": "🟡",
                              "Rilasciato": "🟢", "Obsoleto": "⚫"}.get(
                    drw["state"], "⚪")
                self.lbl_companion.setText(
                    f"✅  DRW {code}  rev.{revision}  —  {state_icon} {drw['state']}"
                )
                self.btn_companion.setVisible(False)
            else:
                self.lbl_companion.setText("⚠️  Nessun DRW associato a questa revisione")
                self.lbl_companion.setStyleSheet("color:#f9e2af;")
                self.btn_companion.setText("＋ Crea DRW")
                self.btn_companion.setVisible(True)
                try:
                    self.btn_companion.clicked.disconnect()
                except Exception:
                    pass
                self.btn_companion.clicked.connect(
                    lambda: self._create_companion("Disegno", doc)
                )

        elif doc_type == "Disegno":
            # Cerca PRT o ASM con stessa revisione
            companion = session.db.fetchone(
                "SELECT * FROM documents WHERE code=? AND doc_type IN ('Parte','Assieme') "
                "AND revision=?",
                (code, revision),
            )
            if companion:
                icon = TYPE_ICON.get(companion["doc_type"], "")
                state_icon = {"In Lavorazione": "🔵", "In Revisione": "🟡",
                              "Rilasciato": "🟢", "Obsoleto": "⚫"}.get(
                    companion["state"], "⚪")
                self.lbl_companion.setText(
                    f"✅  {icon} {companion['doc_type']} {code}  rev.{revision}"
                    f"  —  {state_icon} {companion['state']}"
                )
                self.btn_companion.setVisible(False)
            else:
                self.lbl_companion.setText("⚠️  Nessun PRT/ASM associato a questa revisione")
                self.lbl_companion.setStyleSheet("color:#f9e2af;")
                self.btn_companion.setText("＋ Crea PRT/ASM")
                self.btn_companion.setVisible(True)
                try:
                    self.btn_companion.clicked.disconnect()
                except Exception:
                    pass
                self.btn_companion.clicked.connect(
                    lambda: self._create_companion("PRT_ASM", doc)
                )
        else:
            self.lbl_companion.setText("—")
            self.btn_companion.setVisible(False)

    def _create_companion(self, target_type: str, source_doc: dict):
        """Crea il documento companion (DRW per PRT/ASM o viceversa)."""
        from PyQt6.QtWidgets import QInputDialog
        code     = source_doc["code"]
        revision = source_doc["revision"]
        title    = source_doc.get("title", "")

        if target_type == "Disegno":
            new_type  = "Disegno"
            type_label = "DRW (Disegno)"
        else:
            # Chiedi Parte o Assieme
            choice, ok = QInputDialog.getItem(
                self, "Tipo documento",
                f"Scegli il tipo da creare per {code} rev.{revision}:",
                ["Parte", "Assieme"], 0, False
            )
            if not ok:
                return
            new_type  = choice
            type_label = new_type

        # Titolo pre-compilato, modificabile
        new_title, ok = QInputDialog.getText(
            self, f"Crea {type_label}",
            f"Titolo per {new_type} {code} rev.{revision}:",
            text=title
        )
        if not ok or not new_title.strip():
            return

        # Verifica che non esista già
        if not session.coding.is_code_available(code, revision, new_type):
            QMessageBox.warning(
                self, "Già esistente",
                f"{new_type} {code} rev.{revision} è già presente nel database."
            )
            return

        try:
            parent_doc_id = self.document_id if new_type == "Disegno" else None
            new_id = session.files.create_document(
                code, revision, new_type, new_title.strip(),
                source_doc.get("description", "") or "",
                machine_id=source_doc.get("machine_id"),
                group_id=source_doc.get("group_id"),
                doc_level=source_doc.get("doc_level", 2),
                parent_doc_id=parent_doc_id,
            )
            # R1: allinea lo stato del companion a quello del PRT/ASM
            source_state = source_doc.get("state", "In Lavorazione")
            if source_state != "In Lavorazione":
                session.workflow.sync_companion_state(
                    new_id, source_state, session.user["id"]
                )
            QMessageBox.information(
                self, "OK",
                f"{new_type} creato: {code}  rev.{revision}\n\n"
                f"(id={new_id})"
            )
            # Aggiorna la riga companion
            doc = session.files.get_document(self.document_id)
            if doc:
                self._refresh_companion(doc)
        except Exception as e:
            QMessageBox.critical(self, "Errore", str(e))

    # ------------------------------------------------------------------
    def _refresh_bom(self):
        if not hasattr(self, "tbl_bom"):
            return
        comps = self._collect_bom_rows(self.document_id)
        self.tbl_bom.setRowCount(0)
        self._refresh_bom_commercial()
        for depth, c in comps:
            row = self.tbl_bom.rowCount()
            self.tbl_bom.insertRow(row)
            indent = ("  " * depth) + ("└─ " if depth > 0 else "")
            self.tbl_bom.setItem(row, 0, QTableWidgetItem(indent + c["code"]))
            self.tbl_bom.setItem(row, 1, QTableWidgetItem(c["revision"]))
            self.tbl_bom.setItem(
                row, 2, QTableWidgetItem(
                    TYPE_ICON.get(c["doc_type"], "") + " " + c["doc_type"]
                )
            )
            self.tbl_bom.setItem(row, 3, QTableWidgetItem(c["title"]))
            self.tbl_bom.setItem(row, 4, QTableWidgetItem(str(c["quantity"])))
            self.tbl_bom.item(row, 4).setData(
                Qt.ItemDataRole.UserRole,
                (c["parent_id"], c["child_id"]),
            )

    def _collect_bom_rows(self, parent_id: int,
                          depth: int = 0,
                          visited: Optional[set] = None) -> list[tuple[int, dict]]:
        """Raccoglie la BOM gerarchica in ordine DFS: (depth, componente)."""
        if visited is None:
            visited = set()
        if parent_id in visited:
            return []
        visited.add(parent_id)

        rows: list[tuple[int, dict]] = []
        comps = session.asm.get_components(parent_id)
        for c in comps:
            rows.append((depth, c))
            if c.get("doc_type") == "Assieme":
                rows.extend(self._collect_bom_rows(c["child_id"], depth + 1, visited))
        return rows

    def _refresh_bom_commercial(self):
        """Aggiorna la tabella articoli commerciali nella BOM."""
        if not hasattr(self, "tbl_bom_commercial") or not self.document_id:
            return
        if not session.commercial:
            return
        items = session.commercial.get_commercial_bom(self.document_id)
        self.tbl_bom_commercial.setRowCount(0)
        for it in items:
            r = self.tbl_bom_commercial.rowCount()
            self.tbl_bom_commercial.insertRow(r)
            type_label = "5-COM" if it.get("item_type") == "commerciale" else "6-NOR"
            self.tbl_bom_commercial.setItem(r, 0, QTableWidgetItem(it.get("code") or ""))
            self.tbl_bom_commercial.setItem(r, 1, QTableWidgetItem(type_label))
            self.tbl_bom_commercial.setItem(r, 2, QTableWidgetItem(it.get("description") or ""))
            self.tbl_bom_commercial.setItem(r, 3, QTableWidgetItem(it.get("preferred_supplier") or ""))
            qty_item = QTableWidgetItem(str(it.get("quantity", 1)))
            qty_item.setData(Qt.ItemDataRole.UserRole, it["id"])  # link_id
            self.tbl_bom_commercial.setItem(r, 4, qty_item)

    def _add_commercial_component(self):
        """Aggiunge un articolo commerciale alla BOM dell'assieme."""
        if not self.document_id or not session.commercial:
            return
        from ui.commercial_item_selector import CommercialItemSelectorDialog
        dlg = CommercialItemSelectorDialog(parent=self)
        if dlg.exec() and dlg.selected_id:
            try:
                session.commercial.add_to_bom(
                    self.document_id, dlg.selected_id, quantity=1.0
                )
                self._refresh_bom_commercial()
            except Exception as e:
                QMessageBox.critical(self, "Errore", str(e))

    def _del_commercial_component(self):
        """Rimuove l'articolo commerciale selezionato dalla BOM."""
        row = self.tbl_bom_commercial.currentRow()
        if row < 0:
            return
        qty_item = self.tbl_bom_commercial.item(row, 4)
        if not qty_item:
            return
        link_id = qty_item.data(Qt.ItemDataRole.UserRole)
        if link_id is not None and session.commercial:
            session.commercial.remove_from_bom(link_id)
            self._refresh_bom_commercial()

    def _refresh_history(self):
        if not hasattr(self, "tbl_history"):
            return
        rows = session.workflow.get_history(self.document_id)
        rows += session.checkout.get_log(self.document_id) if session.checkout else []
        rows.sort(key=lambda r: r.get("changed_at") or r.get("timestamp", ""), reverse=True)
        self.tbl_history.setRowCount(0)
        for r in rows:
            i = self.tbl_history.rowCount()
            self.tbl_history.insertRow(i)
            ts     = r.get("changed_at") or r.get("timestamp", "")
            action = r.get("action", "cambio stato")
            frm    = r.get("from_state", "")
            to     = r.get("to_state", "")
            user   = r.get("user_name", "")
            self.tbl_history.setItem(i, 0, QTableWidgetItem(ts[:19]))
            self.tbl_history.setItem(i, 1, QTableWidgetItem(action))
            self.tbl_history.setItem(i, 2, QTableWidgetItem(frm))
            self.tbl_history.setItem(i, 3, QTableWidgetItem(to))
            self.tbl_history.setItem(i, 4, QTableWidgetItem(user))

    # ------------------------------------------------------------------
    # Proprietà
    def _add_property_row(self):
        row = self.tbl_props.rowCount()
        self.tbl_props.insertRow(row)
        self.tbl_props.setItem(row, 0, QTableWidgetItem(""))
        self.tbl_props.setItem(row, 1, QTableWidgetItem(""))
        self.tbl_props.editItem(self.tbl_props.item(row, 0))

    def _remove_property_row(self):
        row = self.tbl_props.currentRow()
        if row >= 0:
            self.tbl_props.removeRow(row)

    def _save_properties(self):
        if not self.document_id:
            return
        props = {}
        for i in range(self.tbl_props.rowCount()):
            name = self.tbl_props.item(i, 0)
            val  = self.tbl_props.item(i, 1)
            if name and name.text().strip():
                props[name.text().strip()] = val.text() if val else ""
        session.properties.save_properties(self.document_id, props)
        QMessageBox.information(self, "OK", f"{len(props)} proprietà salvate")

    def _import_props_from_sw(self):
        if not self.document_id:
            return
        doc = session.files.get_document(self.document_id)
        if not doc or not doc.get("file_name"):
            QMessageBox.warning(self, "Errore", "Nessun file associato al documento")
            return

        # Costruisci path archivio (usato come fallback da read_from_sw_file)
        archive_path = None
        if doc.get("archive_path"):
            archive_path = session.checkout.sp.root / doc["archive_path"]

        try:
            # Import centralizzato: salva custom e applica mapping PDM<->SW.
            sync = session.properties.sync_sw_to_pdm(
                self.document_id,
                archive_path or Path(doc["file_name"]),
                file_name=doc["file_name"],
            )

            if not sync.get("ok"):
                QMessageBox.warning(
                    self, "Errore lettura SolidWorks",
                    f"Impossibile leggere le proprietà da SolidWorks:\n\n{sync.get('error', '')}\n\n"
                    "Assicurarsi che il file sia aperto in SolidWorks.",
                )
                return

            imported_count = int(sync.get("imported_count", 0))
            updated_owner = bool(sync.get("updated_owner", False))

            # Ricarica sempre il documento: titolo/descrizione potrebbero essere stati
            # aggiornati dal mapping anche senza custom properties salvate.
            self._load_document()

            if imported_count <= 0:
                if updated_owner:
                    QMessageBox.information(
                        self, "Importazione completata",
                        "Nessuna proprietà custom trovata nel file SolidWorks,\n"
                        "ma i campi PDM mappati (es. titolo/descrizione) sono stati aggiornati.",
                    )
                    return
                QMessageBox.information(
                    self, "Nessuna proprietà",
                    "Nessuna proprietà custom trovata nel file SolidWorks.",
                )
                return

            # Aggiorna tabella: pulisci e ricarica dal DB
            self.tbl_props.setRowCount(0)
            all_props = session.properties.get_properties(self.document_id)
            for name, value in all_props.items():
                row = self.tbl_props.rowCount()
                self.tbl_props.insertRow(row)
                self.tbl_props.setItem(row, 0, QTableWidgetItem(name))
                self.tbl_props.setItem(row, 1, QTableWidgetItem(value))

            QMessageBox.information(
                self, "Importazione completata",
                f"{imported_count} proprietà importate da SolidWorks.",
            )
        except Exception as e:
            QMessageBox.critical(self, "Errore", str(e))

    def _export_props_excel(self):
        if not self.document_id:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Esporta proprietà", "", "Excel (*.xlsx)"
        )
        if path:
            try:
                session.properties.export_to_excel(
                    self.document_id, Path(path)
                )
                QMessageBox.information(self, "OK", f"Esportato: {path}")
            except Exception as e:
                QMessageBox.critical(self, "Errore", str(e))

    def _import_props_excel(self):
        if not self.document_id:
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Importa proprietà", "", "Excel (*.xlsx)"
        )
        if path:
            try:
                n = session.properties.import_from_excel(
                    self.document_id, Path(path)
                )
                self._load_document()
                QMessageBox.information(self, "OK", f"{n} proprietà importate")
            except Exception as e:
                QMessageBox.critical(self, "Errore", str(e))

    # ------------------------------------------------------------------
    # Gestione file
    def _get_workspace_or_warn(self) -> "Path | None":
        """Ritorna workspace configurata; se assente mostra warning e ritorna None."""
        from ui.sw_config_dialog import SWConfigDialog
        ws = SWConfigDialog.get_workspace()
        if not ws:
            QMessageBox.warning(
                self, "Workspace non configurata",
                "Nessuna workspace configurata su questa workstation.\n"
                "Aprire Strumenti \u2192 Configurazione SolidWorks e impostare la cartella workspace."
            )
        return ws

    def _copy_from_file(self):
        """
        Crea da file: seleziona un file da qualsiasi cartella,
        lo copia in workspace rinominandolo col codice PDM.
        Per PRT/ASM cerca un DRW companion nella stessa cartella.
        """
        if not self.document_id:
            return
        ws = self._get_workspace_or_warn()
        if not ws:
            return
        doc = session.files.get_document(self.document_id)
        # Filtro per tipo documento
        type_filters = {
            "Parte":   "SolidWorks Parte (*.SLDPRT *.sldprt)",
            "Assieme": "SolidWorks Assieme (*.SLDASM *.sldasm)",
            "Disegno": "SolidWorks Disegno (*.SLDDRW *.slddrw)",
        }
        exts_filter = type_filters.get(doc["doc_type"] if doc else "", "File SolidWorks (*.SLDPRT *.SLDASM *.SLDDRW)")
        path, _ = QFileDialog.getOpenFileName(
            self, "Seleziona file sorgente", "", exts_filter
        )
        if not path:
            return
        src_path = Path(path)
        try:
            dest = session.files.copy_to_workspace(src_path, self.document_id, ws)
            QMessageBox.information(
                self, "File copiato in workspace",
                f"File copiato in workspace:\n{dest}\n\n"
                "Aprire il file in SolidWorks, eventualmente aggiornare riferimenti,\n"
                "salvare, poi usare 'Importa in PDM' per archiviarlo."
            )
        except ValueError as e:
            QMessageBox.warning(self, "Tipo file non valido", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Errore", str(e))
            return

        # Per PRT/ASM cerca companion DRW nella stessa cartella
        if doc and doc["doc_type"] in ("Parte", "Assieme"):
            companion = session.files.find_companion_drw(src_path)
            if companion:
                r = QMessageBox.question(
                    self,
                    "Copia DRW associato?",
                    f"Trovato disegno associato:\n{companion.name}\n"
                    "\nCopiarlo in workspace con il codice PDM?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                if r == QMessageBox.StandardButton.Yes:
                    try:
                        drw_doc_id = session.files.get_or_create_drw_document(
                            self.document_id
                        )
                        drw_dest = session.files.copy_to_workspace(
                            companion, drw_doc_id, ws
                        )
                        QMessageBox.information(
                            self, "DRW copiato",
                            f"File DRW copiato in workspace:\n{drw_dest}"
                        )
                    except Exception as e:
                        QMessageBox.critical(self, "Errore DRW", str(e))

    def _import_from_pdm(self):
        """
        Importa in PDM: cerca nella workspace il file con il codice PDM
        e lo archivia. Se non trovato mostra avviso.
        """
        if not self.document_id:
            return
        ws = self._get_workspace_or_warn()
        if not ws:
            return
        doc = session.files.get_document(self.document_id)
        try:
            dest = session.files.import_from_workspace(self.document_id, ws)
            QMessageBox.information(
                self, "Importato in PDM",
                f"File archiviato nel PDM:\n{dest}"
            )
            self._load_document()
        except FileNotFoundError as e:
            QMessageBox.warning(self, "File non trovato in workspace", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Errore", str(e))
            return

        # Per PRT/ASM: cerca DRW in workspace e proponi importazione
        if doc and doc["doc_type"] in ("Parte", "Assieme"):
            try:
                drw_doc_id = session.files.get_or_create_drw_document(
                    self.document_id
                )
                drw_dest = session.files.import_from_workspace(drw_doc_id, ws)
                QMessageBox.information(
                    self, "DRW importato in PDM",
                    f"File DRW archiviato:\n{drw_dest}"
                )
            except FileNotFoundError:
                pass  # DRW non presente in workspace, non è un errore
            except Exception as e:
                QMessageBox.critical(self, "Errore DRW", str(e))

    def _export_file(self):
        """Esporta dall'archivio alla workspace locale."""
        if not self.document_id:
            return
        ws = self._get_workspace_or_warn()
        if not ws:
            return
        doc = session.files.get_document(self.document_id)
        # Proponi DRW solo se il documento è PRT o ASM
        include_drw = False
        if doc and doc["doc_type"] in ("Parte", "Assieme"):
            drw_doc = session.files.get_drw_document(self.document_id)
            if drw_doc and drw_doc.get("archive_path"):
                r = QMessageBox.question(
                    self,
                    "Esporta DRW?",
                    f"Presente disegno associato ({drw_doc['code']}.SLDDRW).\n"
                    "Esportarlo insieme al file principale nella workspace?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                include_drw = (r == QMessageBox.StandardButton.Yes)
        # Pre-check: file già in workspace con contenuto diverso?
        if doc and doc.get("archive_path") and session.sp:
            from pathlib import Path
            from core.file_manager import EXT_FOR_TYPE
            from core.checkout_manager import CheckoutManager
            arch_file = session.sp.root / doc["archive_path"]
            ext = EXT_FOR_TYPE.get(doc.get("doc_type", ""), ".SLDPRT")
            ws_file = Path(ws) / (doc["code"] + ext)
            if ws_file.exists() and arch_file.exists():
                ws_md5   = CheckoutManager._md5(ws_file)
                arch_md5 = CheckoutManager._md5(arch_file)
                if ws_md5 != arch_md5:
                    r = QMessageBox.question(
                        self, "File già in workspace",
                        f"{ws_file.name} è già presente nella workspace\n"
                        f"e differisce dalla versione archiviata.\n\n"
                        f"Sovrascrivere il file locale con la versione dell'archivio?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No,
                    )
                    if r != QMessageBox.StandardButton.Yes:
                        return

        try:
            exported = session.files.export_to_workspace(
                self.document_id, ws, include_drw=include_drw
            )
            if exported:
                msg = "\n".join(str(p) for p in exported)
                QMessageBox.information(
                    self, "File in workspace",
                    f"File copiati nella workspace:\n{msg}"
                )
            else:
                QMessageBox.warning(
                    self, "Nessun file",
                    "Nessun file in archivio da esportare.\n"
                    "Usare prima 'Importa in PDM'."
                )
        except Exception as e:
            QMessageBox.critical(self, "Errore", str(e))

    def _open_sw(self):
        """Copia dall'archivio in workspace (se necessario) e apre in SolidWorks."""
        if not self.document_id:
            return
        ws = self._get_workspace_or_warn()
        if not ws:
            return
        try:
            session.files.open_from_workspace(self.document_id, ws)
        except Exception as e:
            QMessageBox.critical(self, "Errore", str(e))

    def _create_from_template(self):
        """
        Crea il file SW da template tramite SolidWorks COM API.
        Riutilizza la sessione SW attiva se disponibile, altrimenti ne apre una.
        Fallback a copia file se COM non disponibile.
        """
        if not self.document_id:
            return
        ws = self._get_workspace_or_warn()
        if not ws:
            return

        doc = session.files.get_document(self.document_id)
        also_drw = False
        if doc and doc["doc_type"] in ("Parte", "Assieme"):
            from ui.sw_config_dialog import SWConfigDialog
            drw_tpl = SWConfigDialog.get_template("Disegno")
            if drw_tpl and drw_tpl.exists():
                r = QMessageBox.question(
                    self, "Crea DRW?",
                    "Creare anche il file disegno (.SLDDRW) nella workspace?\n"
                    "(verrà creato anche il documento Disegno nel PDM)",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                also_drw = (r == QMessageBox.StandardButton.Yes)

        from config import load_local_config
        from pathlib import Path
        import shutil
        cfg      = load_local_config()
        _EXT     = {"Parte": ".SLDPRT", "Assieme": ".SLDASM", "Disegno": ".SLDDRW"}
        _TPL_KEY = {"Parte": "sw_template_prt", "Assieme": "sw_template_asm",
                    "Disegno": "sw_template_drw"}
        tpl_path = Path(cfg.get(_TPL_KEY.get(doc["doc_type"], ""), "") or "")
        if not tpl_path.exists():
            QMessageBox.warning(
                self, "Template mancante",
                f"Template non configurato per '{doc['doc_type']}'.\n"
                "Configurarlo in Strumenti \u2192 Configurazione SolidWorks."
            )
            return

        dest = Path(ws) / (doc["code"] + _EXT.get(doc["doc_type"], ".SLDPRT"))

        # Connetti a SolidWorks (sessione attiva o nuova)
        sw = None
        try:
            import win32com.client
            try:
                sw = win32com.client.GetActiveObject("SldWorks.Application")
            except Exception:
                sw = win32com.client.Dispatch("SldWorks.Application")
            sw.Visible = True
        except Exception:
            sw = None

        # Crea file principale
        try:
            if sw is not None:
                new_doc = sw.NewDocument(
                    str(tpl_path).replace("/", "\\"), 0, 0, 0
                )
                if new_doc is None:
                    raise RuntimeError(
                        "SolidWorks NewDocument ha restituito None.\n"
                        "Verificare che il file template sia valido e non sia già aperto."
                    )
                new_doc.SaveAs(str(dest).replace("/", "\\"))
            else:
                shutil.copy2(str(tpl_path), str(dest))
        except Exception as e:
            QMessageBox.critical(self, "Errore creazione file", str(e))
            return

        # Crea DRW companion se richiesto
        drw_dest = None
        if also_drw:
            drw_tpl_path = Path(cfg.get("sw_template_drw", "") or "")
            if drw_tpl_path.exists():
                drw_dest = Path(ws) / (doc["code"] + ".SLDDRW")
                try:
                    if sw is not None:
                        drw_doc = sw.NewDocument(
                            str(drw_tpl_path).replace("/", "\\"), 0, 0, 0
                        )
                        if drw_doc:
                            drw_doc.SaveAs(str(drw_dest).replace("/", "\\"))
                        else:
                            shutil.copy2(str(drw_tpl_path), str(drw_dest))
                    else:
                        shutil.copy2(str(drw_tpl_path), str(drw_dest))
                except Exception as e:
                    QMessageBox.warning(self, "Errore DRW", str(e))
                    drw_dest = None
                try:
                    session.files.get_or_create_drw_document(self.document_id)
                except Exception:
                    pass

        msg = f"File creato:\n{dest}"
        if drw_dest:
            msg += f"\n\nDRW creato:\n{drw_dest}"
        if sw is None:
            msg += (
                "\n\n\u26a0\ufe0f SolidWorks non raggiungibile via COM: file copiato dal template.\n"
                "Aprirlo manualmente in SolidWorks."
            )
        QMessageBox.information(self, "File creato", msg)
        if also_drw:
            self._refresh_companion(doc)

    # ------------------------------------------------------------------
    # BOM
    def _add_component(self):
        from ui.document_selector import DocumentSelectorDialog
        dlg = DocumentSelectorDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            child_id = dlg.selected_id
            if child_id and child_id != self.document_id:
                session.asm.add_component(self.document_id, child_id, 1.0)
                self._refresh_bom()

    def _del_component(self):
        row = self.tbl_bom.currentRow()
        if row < 0:
            return
        child_id = self.tbl_bom.item(row, 4)
        if child_id:
            rel = child_id.data(Qt.ItemDataRole.UserRole)
            if isinstance(rel, tuple) and len(rel) == 2:
                pid, cid = rel
            else:
                pid, cid = self.document_id, rel
            session.asm.remove_component(pid, cid)
            self._refresh_bom()

    def _import_asm_from_sw(self):
        if not self.document_id:
            return
        doc = session.files.get_document(self.document_id)
        if not doc or doc.get("doc_type") != "Assieme":
            QMessageBox.warning(
                self, "Errore",
                "Questa funzione è disponibile solo per documenti di tipo Assieme."
            )
            return
        if not doc.get("archive_path") or not session.sp:
            QMessageBox.warning(
                self, "File non archiviato",
                "Il file non è ancora archiviato nel PDM.\n"
                "Eseguire il check-in prima di importare la struttura."
            )
            return
        asm_file = session.sp.root / doc["archive_path"]
        if not asm_file.exists():
            QMessageBox.warning(
                self, "File non trovato",
                f"Il file archiviato non è stato trovato:\n{asm_file}"
            )
            return
        try:
            n = session.asm.import_from_sw_asm(asm_file, self.document_id)
            self._refresh_bom()
            QMessageBox.information(
                self, "OK",
                f"Importati {n} componenti dall'assieme.\n\n"
                "Nota: vengono collegati solo i documenti già presenti nel PDM\n"
                "(codice = nome file senza estensione)."
            )
        except Exception as e:
            QMessageBox.critical(self, "Errore import BOM", str(e))

    def _export_bom_excel(self):
        if not self.document_id:
            return
        components = session.asm.get_bom_flat(self.document_id)
        if not components:
            QMessageBox.information(self, "BOM vuota", "Nessun componente nella BOM.")
            return

        doc = session.files.get_document(self.document_id)
        default_name = f"BOM_{doc['code']}_{doc['revision']}.xlsx" if doc else "BOM.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self, "Esporta BOM Excel", default_name,
            "Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM"

            if doc:
                ws.append([f"BOM - {doc['code']} rev.{doc['revision']} - {doc['title']}"])
                ws["A1"].font = Font(bold=True, size=12)
                ws.append([])

            headers = ["#", "Codice", "Rev.", "Tipo", "Titolo", "Qta"]
            ws.append(headers)
            hdr_row = ws.max_row
            fill = PatternFill("solid", fgColor="1E3A5F")
            font_hdr = Font(bold=True, color="FFFFFF")
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(hdr_row, col)
                cell.fill = fill
                cell.font = font_hdr
                cell.alignment = Alignment(horizontal="center")

            for i, comp in enumerate(components, 1):
                ws.append([
                    i,
                    comp.get("code", ""),
                    comp.get("revision", ""),
                    comp.get("doc_type", ""),
                    comp.get("title", ""),
                    comp.get("quantity", 1),
                ])

            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 6
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 40
            ws.column_dimensions["F"].width = 8

            wb.save(path)
            QMessageBox.information(
                self, "Esportazione completata",
                f"BOM esportata in:\n{path}\n\n{len(components)} componenti."
            )
        except Exception as e:
            QMessageBox.critical(self, "Errore export BOM Excel", str(e))
