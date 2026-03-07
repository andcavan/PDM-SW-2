# =============================================================================
#  core/properties_manager.py  –  Import/export proprietà SolidWorks
# =============================================================================
from __future__ import annotations
import re
from pathlib import Path
from typing import Optional, TYPE_CHECKING

from config import load_local_config

if TYPE_CHECKING:
    from core.database import Database


# Proprietà standard SolidWorks
SW_STANDARD_PROPS = [
    "Description", "PartNo", "DrawnBy", "DrawnDate",
    "CheckedBy", "CheckedDate", "EngineeringApproval",
    "ManufacturingApproval", "QualityApproval",
    "Material", "Finish", "Weight", "Density",
    "Revision", "Title", "Project", "Company",
]


class PropertiesManager:
    MAPPABLE_FIELDS = (
        "revision",
        "title",
        "description",
        "code",
        "state",
        "created_by",
        "created_at",
    )

    def __init__(self, db: "Database"):
        self.db = db

    # ------------------------------------------------------------------
    # DB
    # ------------------------------------------------------------------
    def save_properties(self, document_id: int, props: dict):
        """Salva un dizionario {nome: valore} nel DB."""
        for name, value in props.items():
            existing = self.db.fetchone(
                "SELECT id FROM document_properties WHERE document_id=? AND prop_name=?",
                (document_id, name),
            )
            if existing:
                self.db.execute(
                    "UPDATE document_properties SET prop_value=? WHERE id=?",
                    (str(value), existing["id"]),
                )
            else:
                self.db.execute(
                    """INSERT INTO document_properties (document_id, prop_name, prop_value)
                       VALUES (?,?,?)""",
                    (document_id, name, str(value)),
                )

    def get_properties(self, document_id: int) -> dict:
        rows = self.db.fetchall(
            "SELECT prop_name, prop_value FROM document_properties WHERE document_id=?",
            (document_id,),
        )
        return {r["prop_name"]: r["prop_value"] for r in rows}

    def delete_property(self, document_id: int, prop_name: str):
        self.db.execute(
            "DELETE FROM document_properties WHERE document_id=? AND prop_name=?",
            (document_id, prop_name),
        )

    @staticmethod
    def _default_property_mapping() -> dict:
        return {
            "revision": {
                "sw_names": ["Revision", "Rev"],
                "mode": "PDM->SW",
                "force_pdm": True,
            },
            "title": {
                "sw_names": ["Title", "Titolo", "Subject"],
                "mode": "Bidirezionale",
                "force_pdm": False,
            },
            "description": {
                "sw_names": ["Description", "Descrizione", "Comments", "Subject"],
                "mode": "Bidirezionale",
                "force_pdm": False,
            },
            "code": {
                "sw_names": ["PartNo", "Code", "Codice"],
                "mode": "PDM->SW",
                "force_pdm": True,
            },
            "state": {
                "sw_names": ["State", "Stato"],
                "mode": "PDM->SW",
                "force_pdm": True,
            },
            "created_by": {
                "sw_names": ["DrawnBy", "CreatedBy", "CreatoDa", "Author"],
                "mode": "PDM->SW",
                "force_pdm": True,
            },
            "created_at": {
                "sw_names": ["DrawnDate", "CreatedDate", "DataCreazione", "Date"],
                "mode": "PDM->SW",
                "force_pdm": True,
            },
        }

    @staticmethod
    def _normalize_mapping(raw) -> dict:
        default = PropertiesManager._default_property_mapping()
        out: dict = {}
        raw = raw if isinstance(raw, dict) else {}
        for fld in PropertiesManager.MAPPABLE_FIELDS:
            item = raw.get(fld) if isinstance(raw.get(fld), dict) else {}
            names = item.get("sw_names", default[fld]["sw_names"])
            if not isinstance(names, list):
                names = [str(names)] if names else []
            names_clean: list[str] = []
            for n in names:
                s = str(n).strip()
                if s and s not in names_clean:
                    names_clean.append(s)
            if not names_clean:
                names_clean = list(default[fld]["sw_names"])
            mode = str(item.get("mode") or default[fld]["mode"])
            if mode not in ("SW->PDM", "PDM->SW", "Bidirezionale"):
                mode = default[fld]["mode"]
            out[fld] = {
                "sw_names": names_clean,
                "mode": mode,
                "force_pdm": bool(item.get("force_pdm", default[fld]["force_pdm"])),
            }
        return out

    def get_property_mapping(self) -> dict:
        cfg = load_local_config()
        raw = cfg.get("sw_property_mapping")
        return self._normalize_mapping(raw)

    @staticmethod
    def _mode_allows_sw_to_pdm(mode: str) -> bool:
        return mode in ("SW->PDM", "Bidirezionale")

    @staticmethod
    def _mode_allows_pdm_to_sw(mode: str) -> bool:
        return mode in ("PDM->SW", "Bidirezionale")

    @staticmethod
    def _normalize_prop_key(name: str) -> str:
        return re.sub(r"[^A-Z0-9]", "", str(name or "").strip().upper())

    @staticmethod
    def _pick_first_prop_value(props: dict, names: list[str]) -> str:
        if not isinstance(props, dict):
            return ""
        by_key: dict[str, str] = {}
        for k, v in props.items():
            nk = PropertiesManager._normalize_prop_key(k)
            if nk and nk not in by_key:
                by_key[nk] = str(v or "")
        for name in names:
            key = PropertiesManager._normalize_prop_key(name)
            if key and key in by_key:
                val = by_key[key].strip()
                if val:
                    return val
        return ""

    @staticmethod
    def _collapse_alias_properties(props: dict) -> dict:
        """Unifica alias bilingua per evitare doppioni nel DB proprieta."""
        out = {str(k): str(v or "") for k, v in (props or {}).items()}
        alias_groups = [
            ("Title", ["Title", "Titolo"]),
            ("Description", ["Description", "Descrizione"]),
            ("Author", ["Author", "Autore"]),
        ]

        # Mappa chiavi normalizzate -> chiave reale presente in out
        key_by_norm: dict[str, str] = {}
        for k in list(out.keys()):
            n = PropertiesManager._normalize_prop_key(k)
            if n and n not in key_by_norm:
                key_by_norm[n] = k

        for canonical, aliases in alias_groups:
            value = PropertiesManager._pick_first_prop_value(out, aliases)
            if not value:
                continue

            out[canonical] = value
            for a in aliases:
                an = PropertiesManager._normalize_prop_key(a)
                old_key = key_by_norm.get(an)
                if old_key and old_key in out and old_key != canonical:
                    del out[old_key]

        return out

    def resolve_property_owner_doc(self, document_id: int) -> int:
        """
        Ritorna il documento owner per i campi PDM fondamentali.
        Regola: i DRW ereditano da PRT/ASM padre.
        """
        doc = self.db.fetchone(
            "SELECT id, code, revision, doc_type, parent_doc_id FROM documents WHERE id=?",
            (document_id,),
        )
        if not doc:
            raise ValueError("Documento non trovato")

        if doc.get("doc_type") != "Disegno":
            return int(doc["id"])

        parent_id = doc.get("parent_doc_id")
        if parent_id:
            parent = self.db.fetchone(
                "SELECT id FROM documents WHERE id=? AND doc_type IN ('Parte','Assieme')",
                (parent_id,),
            )
            if parent:
                return int(parent["id"])

        parent = self.db.fetchone(
            "SELECT id FROM documents "
            "WHERE code=? AND revision=? AND doc_type IN ('Parte','Assieme') "
            "ORDER BY id DESC LIMIT 1",
            (doc["code"], doc["revision"]),
        )
        if parent:
            return int(parent["id"])

        return int(doc["id"])

    def sync_sw_to_pdm(self, document_id: int, file_path: Path,
                       file_name: str | None = None) -> dict:
        """
        Importa proprieta da SW nel DB:
        - salva sempre tutte le custom in document_properties del documento corrente
        - aggiorna title/description dell'owner PDM secondo mappatura
        - non aggiorna la revisione da SW (source of truth: PDM)
        """
        props = self.read_from_sw_file(file_path, file_name=file_name)
        err = props.pop("_error", None)
        if err:
            return {"ok": False, "error": str(err), "imported_count": 0, "updated_owner": False}

        props_to_save = self._collapse_alias_properties(props)
        if props_to_save:
            self.save_properties(document_id, props_to_save)

        mapping = self.get_property_mapping()
        owner_id = self.resolve_property_owner_doc(document_id)
        owner = self.db.fetchone(
            "SELECT id, title, description FROM documents WHERE id=?",
            (owner_id,),
        )

        new_title = owner.get("title") if owner else ""
        new_desc = owner.get("description") if owner else ""

        title_map = mapping.get("title", {})
        if self._mode_allows_sw_to_pdm(str(title_map.get("mode", ""))):
            val = self._pick_first_prop_value(props, list(title_map.get("sw_names", [])))
            if val:
                new_title = val

        desc_map = mapping.get("description", {})
        if self._mode_allows_sw_to_pdm(str(desc_map.get("mode", ""))):
            val = self._pick_first_prop_value(props, list(desc_map.get("sw_names", [])))
            if val:
                new_desc = val

        updated_owner = False
        if owner and (str(new_title or "") != str(owner.get("title") or "") or
                      str(new_desc or "") != str(owner.get("description") or "")):
            self.db.execute(
                "UPDATE documents SET title=?, description=?, modified_at=datetime('now') WHERE id=?",
                (str(new_title or ""), str(new_desc or ""), owner_id),
            )
            updated_owner = True

        return {
            "ok": True,
            "error": "",
            "imported_count": len(props_to_save),
            "updated_owner": updated_owner,
            "owner_id": owner_id,
            "props": props_to_save,
        }

    def sync_pdm_to_sw(self, document_id: int, file_path: Path,
                       force_revision: bool = True) -> dict:
        """
        Esporta campi PDM fondamentali verso SW secondo mappatura.
        Per i DRW usa come owner PDM il PRT/ASM padre.
        """
        mapping = self.get_property_mapping()
        owner_id = self.resolve_property_owner_doc(document_id)
        owner = self.db.fetchone(
            "SELECT id, code, revision, title, description, state, created_by, created_at "
            "FROM documents WHERE id=?",
            (owner_id,),
        )
        if not owner:
            return {"ok": False, "error": "Documento owner non trovato", "written_count": 0}

        out_props: dict[str, str] = {}

        created_by_name = ""
        created_by_id = owner.get("created_by")
        if created_by_id:
            user = self.db.fetchone(
                "SELECT full_name, username FROM users WHERE id=?",
                (created_by_id,),
            )
            if user:
                created_by_name = str(user.get("full_name") or user.get("username") or "").strip()

        rev_map = mapping.get("revision", {})
        rev_names = list(rev_map.get("sw_names", []))
        rev_mode = str(rev_map.get("mode", ""))
        rev_force = bool(rev_map.get("force_pdm", False)) or bool(force_revision)
        if rev_names and rev_force and self._mode_allows_pdm_to_sw(rev_mode):
            out_props[rev_names[0]] = str(owner.get("revision") or "")

        title_map = mapping.get("title", {})
        title_names = list(title_map.get("sw_names", []))
        if title_names and self._mode_allows_pdm_to_sw(str(title_map.get("mode", ""))):
            out_props[title_names[0]] = str(owner.get("title") or "")

        desc_map = mapping.get("description", {})
        desc_names = list(desc_map.get("sw_names", []))
        if desc_names and self._mode_allows_pdm_to_sw(str(desc_map.get("mode", ""))):
            out_props[desc_names[0]] = str(owner.get("description") or "")

        code_map = mapping.get("code", {})
        code_names = list(code_map.get("sw_names", []))
        if code_names and self._mode_allows_pdm_to_sw(str(code_map.get("mode", ""))):
            out_props[code_names[0]] = str(owner.get("code") or "")

        state_map = mapping.get("state", {})
        state_names = list(state_map.get("sw_names", []))
        if state_names and self._mode_allows_pdm_to_sw(str(state_map.get("mode", ""))):
            out_props[state_names[0]] = str(owner.get("state") or "")

        created_by_map = mapping.get("created_by", {})
        created_by_names = list(created_by_map.get("sw_names", []))
        if created_by_names and self._mode_allows_pdm_to_sw(str(created_by_map.get("mode", ""))):
            out_props[created_by_names[0]] = created_by_name

        created_at_map = mapping.get("created_at", {})
        created_at_names = list(created_at_map.get("sw_names", []))
        if created_at_names and self._mode_allows_pdm_to_sw(str(created_at_map.get("mode", ""))):
            out_props[created_at_names[0]] = str(owner.get("created_at") or "")

        if not out_props:
            return {"ok": True, "error": "", "written_count": 0, "owner_id": owner_id}

        self.write_to_sw_file(file_path, out_props)
        return {
            "ok": True,
            "error": "",
            "written_count": len(out_props),
            "owner_id": owner_id,
            "props": out_props,
        }

    # ------------------------------------------------------------------
    # COM – SolidWorks API (richiede pywin32 e SW installato)
    # ------------------------------------------------------------------
    def read_from_sw_file(self, file_path: Path, file_name: str = None) -> dict:
        """
        Legge le proprietà da un file SolidWorks.
        Se SW è già in esecuzione e il file è aperto, usa il doc esistente.
        Cerca per: 1) path esatto, 2) file_name tra i doc aperti, 3) OpenDoc6.
        """
        try:
            import win32com.client as win32
        except ImportError:
            raise ImportError(
                "pywin32 non installato. "
                "Eseguire: pip install pywin32"
            )

        props = {}
        opened_by_us = False
        sw = None
        model = None

        try:
            # Prova a connettersi all'istanza SW già in esecuzione
            try:
                sw = win32.GetActiveObject("SldWorks.Application")
            except Exception:
                sw = win32.Dispatch("SldWorks.Application")

            if sw is None:
                props["_error"] = "Impossibile connettersi a SolidWorks"
                return props

            search_name = (file_name or file_path.name).upper()

            # 1) Prova il documento attivo (caso più comune)
            #    L'utente ha premuto "Importa da SW" → vuole le props
            #    del doc attivo in SolidWorks, indipendentemente dal nome.
            try:
                active = sw.ActiveDoc
                if active is not None:
                    model = active
            except Exception:
                pass

            # 2) Cerca per path esatto
            if model is None:
                file_str = str(file_path)
                try:
                    model = sw.GetOpenDocumentByName(file_str)
                except Exception:
                    pass

            # 3) Cerca per nome file tra i documenti aperti
            #    Itera i doc, trova il path completo, poi usa
            #    GetOpenDocumentByName per ottenere un dispatch COM corretto
            if model is None:
                found_path = None
                try:
                    doc_iter = sw.GetFirstDocument()
                    while doc_iter is not None:
                        try:
                            fp = doc_iter.GetPathName
                            if isinstance(fp, str) and fp:
                                if Path(fp).name.upper() == search_name:
                                    found_path = fp
                                    break
                        except Exception:
                            pass
                        try:
                            doc_iter = doc_iter.GetNext
                            if doc_iter is None or not hasattr(doc_iter, 'GetPathName'):
                                break
                        except Exception:
                            break
                except Exception:
                    pass

                if found_path:
                    try:
                        model = sw.GetOpenDocumentByName(found_path)
                    except Exception:
                        pass

            # 4) Prova ad aprire il file dal path (fallback)
            if model is None:
                ext = file_path.suffix.upper()
                doc_type_map = {
                    ".SLDPRT": 1,   # swDocPART
                    ".SLDASM": 2,   # swDocASSEMBLY
                    ".SLDDRW": 3,   # swDocDRAWING
                }
                doc_type = doc_type_map.get(ext, 1)
                arg_err = win32.VARIANT(win32.pythoncom.VT_BYREF | win32.pythoncom.VT_I4, 0)
                arg_warn = win32.VARIANT(win32.pythoncom.VT_BYREF | win32.pythoncom.VT_I4, 0)
                model = sw.OpenDoc6(
                    str(file_path), doc_type,
                    1,   # swOpenDocOptions_Silent
                    "", arg_err, arg_warn
                )
                if model:
                    opened_by_us = True

            if model is None:
                props["_error"] = (
                    f"File '{search_name}' non trovato in SolidWorks.\n"
                    "Aprire il file in SolidWorks prima di importare."
                )
                return props

            # Leggi proprietà custom (configurazione default "")
            props = self._read_custom_props(model)

            # Arricchisci con SummaryInfo standard quando disponibili.
            summary_fields = {
                "Title": "Title",
                "Subject": "Subject",
                "Comments": "Comments",
                "Author": "Author",
                "CreatedDate": "CreatedDate",
                "SaveDate": "SaveDate",
            }
            for prop_name, summary_key in summary_fields.items():
                if prop_name not in props:
                    val = self._read_summary_info(model, summary_key)
                    if val:
                        props[prop_name] = val

            # Compatibilità mapping: Description può provenire da Subject/Comments.
            if "Description" not in props:
                props["Description"] = str(props.get("Subject") or props.get("Comments") or "").strip()

        except Exception as e:
            props["_error"] = str(e)
        finally:
            if opened_by_us and model and sw:
                try:
                    sw.CloseDoc(model.GetTitle())
                except Exception:
                    pass
        return props

    @staticmethod
    def _is_link_expr(text: str) -> bool:
        """True se il testo sembra una espressione linkata SW ($PRP/$PRPSHEET)."""
        if not isinstance(text, str):
            return False
        t = text.strip().upper()
        return t.startswith("$PRP:") or t.startswith("$PRPSHEET:")

    @staticmethod
    def _best_value_from_result(result, fallback_expr: str = "") -> str:
        """
        Estrae il valore migliore da un risultato COM, preferendo il valore risolto.
        Molte API SW ritornano tuple con piu campi (raw + resolved + flags).
        """
        strings: list[str] = []
        if isinstance(result, str):
            strings = [result]
        elif isinstance(result, (tuple, list)):
            strings = [str(x) for x in result if isinstance(x, str)]

        if not strings:
            return fallback_expr or ""

        # Cerca prima un valore "risolto" non vuoto e non espressione.
        for s in strings:
            st = s.strip()
            if st and not PropertiesManager._is_link_expr(st):
                return st

        # Altrimenti ritorna il primo testo disponibile (tipicamente espressione raw).
        return strings[0].strip()

    @staticmethod
    def _read_prop_value(mgr, name: str) -> str:
        """
        Legge una custom property privilegiando il valore valutato (resolved).
        Supporta varianti API COM SW: Get6/Get5/Get4/Get2/Get.
        """
        # Tentativo principale: Get6/Get5 con parametri byref (pywin32).
        # Questo e il metodo piu affidabile per ottenere il valore valutato
        # (es. token SW-* come SW-Mass@...).
        try:
            import pythoncom
            import win32com.client as win32
        except Exception:
            pythoncom = None
            win32 = None

        if pythoncom is not None and win32 is not None:
            try:
                val_out = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BSTR, "")
                res_out = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BSTR, "")
                was_res = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BOOL, False)
                link_to = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BOOL, False)
                _ = mgr.Get6(name, False, val_out, res_out, was_res, link_to)
                resolved = str(res_out.value or "").strip()
                raw_expr = str(val_out.value or "").strip()
                if resolved and not PropertiesManager._is_link_expr(resolved):
                    return resolved
                if raw_expr:
                    return raw_expr
            except Exception:
                pass

            try:
                val_out = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BSTR, "")
                res_out = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BSTR, "")
                was_res = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BOOL, False)
                _ = mgr.Get5(name, False, val_out, res_out, was_res)
                resolved = str(res_out.value or "").strip()
                raw_expr = str(val_out.value or "").strip()
                if resolved and not PropertiesManager._is_link_expr(resolved):
                    return resolved
                if raw_expr:
                    return raw_expr
            except Exception:
                pass

        # API recenti: Get6/Get5 espongono raw + resolved.
        for method_name, args in (
            ("Get6", (name, False, "", "", False, False)),
            ("Get5", (name, False, "", "", False)),
            ("Get4", (name, False, "", "")),
            ("Get2", (name, "", "")),
        ):
            try:
                method = getattr(mgr, method_name)
                res = method(*args)
                val = PropertiesManager._best_value_from_result(res)
                if val != "":
                    return val
            except Exception:
                continue

        # Fallback legacy: Get(name) (spesso raw expression)
        try:
            raw = mgr.Get(name)
            return PropertiesManager._best_value_from_result(raw)
        except Exception:
            return ""

    @staticmethod
    def _read_summary_info(model, field: str) -> str:
        """Legge un campo SummaryInfo SW usando alias noti (es. SW-Author)."""
        raw_key = (field or "").strip()
        # Gestisce target con suffisso configurazione/file, es. "Description@Part1.SLDPRT"
        key = raw_key.split("@", 1)[0].strip().upper().replace("_", "-")
        key_compact = key.replace(" ", "").replace("-", "")

        # Campi "speciali" legati al file (non SummaryInfo standard)
        file_aliases = {
            "FILENAME", "SWFILENAME", "SWFILENAMEEXT", "SWFILE", "SWFILENAMEWITHEXTENSION",
        }
        file_noext_aliases = {
            "FILENAMEWITHOUTEXTENSION", "FILENAMEWITHEXTENSIONREMOVED", "SWFILENAMEWITHOUTEXTENSION",
        }

        model_path = ""
        try:
            p = model.GetPathName
            p = p() if callable(p) else p
            model_path = str(p or "").strip()
        except Exception:
            model_path = ""

        if key_compact in file_aliases:
            if model_path:
                return Path(model_path).name
            try:
                t = model.GetTitle
                t = t() if callable(t) else t
                return str(t or "").strip()
            except Exception:
                return ""

        if key_compact in file_noext_aliases:
            if model_path:
                return Path(model_path).stem
            try:
                t = model.GetTitle
                t = t() if callable(t) else t
                return Path(str(t or "").strip()).stem
            except Exception:
                return ""

        # swSummInfoField_e principali
        idx_map = {
            "TITLE": 0,
            "SUBJECT": 1,
            "AUTHOR": 2,
            "KEYWORDS": 3,
            "COMMENTS": 4,
            "SAVEDBY": 5,
            "CREATEDDATE": 6,
            "SAVEDATE": 7,
        }

        # Accetta sia alias standard (Author) che SW-* (SW-Author)
        if key_compact.startswith("SW"):
            key_compact = key_compact[2:]
        idx = idx_map.get(key_compact)
        if idx is None:
            return ""
        try:
            val = model.SummaryInfo(idx)
            return str(val).strip() if val else ""
        except Exception:
            return ""

    @staticmethod
    def _parse_link_expression(expr: str) -> tuple[str, str] | tuple[None, None]:
        """Parsa link SW e ritorna (kind, target): kind in {PRP, PRPSHEET}."""
        if not isinstance(expr, str):
            return None, None
        m = re.match(
            r'^\s*\$(PRP|PRPSHEET)\s*:\s*(?:"([^"]+)"|\'([^\']+)\'|(.+))\s*$',
            expr,
            re.IGNORECASE,
        )
        if not m:
            return None, None
        kind = (m.group(1) or "").upper()
        target = (m.group(2) or m.group(3) or m.group(4) or "").strip()
        if target.startswith('"') and target.endswith('"') and len(target) >= 2:
            target = target[1:-1].strip()
        if not target:
            return None, None
        return kind, target

    @staticmethod
    def _parse_sw_token_expression(expr: str) -> str:
        """
        Estrae il token SW-* da stringhe tipo:
        - SW-Mass@FILE.SLDASM
        - "SW-Mass@FILE.SLDASM"
        Restituisce il target senza quote, oppure stringa vuota.
        """
        if not isinstance(expr, str):
            return ""
        t = expr.strip()
        if len(t) >= 2 and t[0] == '"' and t[-1] == '"':
            t = t[1:-1].strip()
        if not t.upper().startswith("SW-"):
            return ""
        return t

    @staticmethod
    def _resolve_sw_token_expression(model, expr: str) -> str:
        """
        Risolve token SW-* comuni non racchiusi in $PRP, p.es. "SW-Mass@...".
        """
        token = PropertiesManager._parse_sw_token_expression(expr)
        if not token:
            return ""

        # Token base senza suffisso @config/file
        base = token.split("@", 1)[0].strip().upper().replace("_", "-")
        base_compact = base.replace(" ", "").replace("-", "")

        # Campi summary/file (Author, Title, File Name, ...)
        summary_val = PropertiesManager._read_summary_info(model, token)
        if summary_val:
            return summary_val

        # Proprieta fisiche calcolate
        try:
            ext = model.Extension
            mp = ext.CreateMassProperty() if ext else None
        except Exception:
            mp = None

        if mp is None:
            return ""

        try:
            if base_compact in ("SWMASS",):
                return str(mp.Mass)
            if base_compact in ("SWVOLUME",):
                return str(mp.Volume)
            if base_compact in ("SWDENSITY",):
                return str(mp.Density)
            if base_compact in ("SWSURFACEAREA",):
                return str(mp.SurfaceArea)
        except Exception:
            return ""

        return ""

    @staticmethod
    def _resolve_link_expression(model, mgr, expr: str,
                                 all_mgrs: list | None = None,
                                 depth: int = 0) -> str:
        """
        Risolve espressioni linkate SW, p.es. $PRP:"SW-Author".
        depth evita ricorsioni infinite su catene di proprietà.
        """
        if depth > 2 or not isinstance(expr, str):
            return ""

        kind, target = PropertiesManager._parse_link_expression(expr)
        if not kind:
            return ""

        # 1) Target su SummaryInfo (caso tipico SW-Author, SW-Title...)
        sum_val = PropertiesManager._read_summary_info(model, target)
        if sum_val:
            return sum_val

        # 2) Target come altra custom property (anche con riferimento @config)
        prop_name = target.split("@", 1)[0].strip() or target

        managers: list = []
        if mgr is not None:
            managers.append(mgr)
        if all_mgrs:
            managers.extend(all_mgrs)

        # PRPSHEET spesso punta a proprietà non-file-level: inverti priorità
        if kind == "PRPSHEET" and len(managers) > 1:
            managers = list(reversed(managers))

        # Deduplica manager preservando ordine
        uniq_mgrs: list = []
        for mobj in managers:
            if mobj is None:
                continue
            if all(id(mobj) != id(x) for x in uniq_mgrs):
                uniq_mgrs.append(mobj)

        for mobj in uniq_mgrs:
            linked_val = PropertiesManager._read_prop_value(mobj, prop_name)
            if linked_val and not PropertiesManager._is_link_expr(linked_val):
                return linked_val
            if linked_val and PropertiesManager._is_link_expr(linked_val):
                res = PropertiesManager._resolve_link_expression(
                    model, mobj, linked_val, all_mgrs=uniq_mgrs, depth=depth + 1
                )
                if res:
                    return res

        return ""

    @staticmethod
    def _read_custom_props(model) -> dict:
        """
        Legge le proprietà custom dal modello SW.
        Legge sia dalla config "" (file-level) che dalle configurazioni specifiche.
        Nota: con GetActiveObject, i metodi COM possono essere esposti come
        proprietà (senza parentesi): GetNames → proprietà, Get → proprietà.
        """
        props = {}

        ext = model.Extension

        # Raccogli lista configurazioni da leggere: "" (file-level) + tutte le config
        config_names = [""]
        try:
            cfgs = model.GetConfigurationNames
            if cfgs:
                if isinstance(cfgs, str):
                    config_names.append(cfgs)
                else:
                    config_names.extend(cfgs)
        except Exception:
            pass

        # Pre-carica i manager disponibili per poter risolvere link cross-config.
        mgr_by_cfg: dict[str, object] = {}
        for cfg_name in config_names:
            try:
                m = ext.CustomPropertyManager(cfg_name)
                if isinstance(m, tuple):
                    m = m[0] if m else None
                if m is not None:
                    mgr_by_cfg[cfg_name] = m
            except Exception:
                continue
        all_mgrs = list(mgr_by_cfg.values())

        for cfg_name in config_names:
            mgr = mgr_by_cfg.get(cfg_name)
            if mgr is None:
                continue

            try:
                count = mgr.Count
                if callable(count):
                    count = count()
                if isinstance(count, (tuple, list)):
                    count = count[0] if count else 0
                count = int(count or 0)
                if not count or count <= 0:
                    continue
            except Exception:
                continue

            # GetNames è una proprietà (non metodo) con GetActiveObject
            names = None
            try:
                names = mgr.GetNames
                if callable(names):
                    names = names()
            except Exception:
                pass
            if names is None:
                try:
                    names = mgr.GetNames()
                except Exception:
                    pass
            if callable(names):
                try:
                    names = names()
                except Exception:
                    names = None
            if isinstance(names, tuple):
                names = [x for x in names if isinstance(x, str)]
            if not names:
                continue

            # Se è una stringa singola, wrappala in lista
            if isinstance(names, str):
                names = [names]

            for name in names:
                if name in props:
                    continue  # file-level ha priorità
                try:
                    value = PropertiesManager._read_prop_value(mgr, str(name))
                    if PropertiesManager._is_link_expr(value):
                        resolved = PropertiesManager._resolve_link_expression(
                            model, mgr, value, all_mgrs=all_mgrs
                        )
                        if resolved:
                            value = resolved
                    else:
                        resolved_sw = PropertiesManager._resolve_sw_token_expression(model, value)
                        if resolved_sw:
                            value = resolved_sw
                    props[name] = value
                except Exception:
                    props[name] = ""

        return props

    def write_to_sw_file(self, file_path: Path, props: dict):
        """Scrive le proprietà in un file SolidWorks via COM."""
        try:
            import win32com.client as win32
        except ImportError:
            raise ImportError("pywin32 non installato")

        opened_by_us = False
        sw = None
        model = None

        try:
            # Prova a connettersi all'istanza SW già in esecuzione
            try:
                sw = win32.GetActiveObject("SldWorks.Application")
            except Exception:
                sw = win32.Dispatch("SldWorks.Application")

            if sw is None:
                raise RuntimeError("Impossibile connettersi a SolidWorks")

            # Cerca se il file è già aperto
            file_str = str(file_path)
            model = sw.GetOpenDocumentByName(file_str)

            if model is None:
                # Fallback per file aperto con path differente: match per nome file.
                search_name = file_path.name.upper()
                found_path = None
                try:
                    doc_iter = sw.GetFirstDocument()
                    while doc_iter is not None:
                        try:
                            fp = doc_iter.GetPathName
                            fp = fp() if callable(fp) else fp
                            if isinstance(fp, str) and fp:
                                if Path(fp).name.upper() == search_name:
                                    found_path = fp
                                    break
                        except Exception:
                            pass
                        try:
                            doc_iter = doc_iter.GetNext
                            doc_iter = doc_iter() if callable(doc_iter) else doc_iter
                        except Exception:
                            break
                except Exception:
                    pass
                if found_path:
                    try:
                        model = sw.GetOpenDocumentByName(found_path)
                    except Exception:
                        pass

            if model is None:
                ext = file_path.suffix.upper()
                doc_type_map = {
                    ".SLDPRT": 1, ".SLDASM": 2, ".SLDDRW": 3
                }
                doc_type = doc_type_map.get(ext, 1)
                arg_err = win32.VARIANT(win32.pythoncom.VT_BYREF | win32.pythoncom.VT_I4, 0)
                arg_warn = win32.VARIANT(win32.pythoncom.VT_BYREF | win32.pythoncom.VT_I4, 0)
                model = sw.OpenDoc6(
                    file_str, doc_type,
                    0, "", arg_err, arg_warn
                )
                if model:
                    opened_by_us = True

            if model is None:
                raise RuntimeError(f"Impossibile aprire: {file_path.name}")

            mgr = model.Extension.CustomPropertyManager("")
            for name, value in props.items():
                n = str(name or "").strip()
                if not n:
                    continue
                v = str(value or "")
                # Prima prova Set2 su proprieta esistente.
                try:
                    mgr.Set2(n, v)
                    continue
                except Exception:
                    pass
                # Poi Add3 per nuova proprieta.
                try:
                    # swCustomInfoText=30, swCustomPropertyReplaceValue=2
                    mgr.Add3(n, 30, v, 2)
                except Exception:
                    # Fallback legacy
                    mgr.Add2(n, 30, v)

            try:
                model.Save3(1, 0, 0)
            except Exception:
                model.Save()

        finally:
            if opened_by_us and model and sw:
                try:
                    sw.CloseDoc(model.GetTitle())
                except Exception:
                    pass

    # ------------------------------------------------------------------
    # Export Excel
    # ------------------------------------------------------------------
    def export_to_excel(self, document_id: int, dest: Path):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl non installato")

        props = self.get_properties(document_id)
        doc   = self.db.fetchone(
            "SELECT code, revision, title FROM documents WHERE id=?",
            (document_id,),
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proprietà"
        ws.append(["Codice", doc["code"] if doc else ""])
        ws.append(["Revisione", doc["revision"] if doc else ""])
        ws.append(["Titolo", doc["title"] if doc else ""])
        ws.append([])
        ws.append(["Proprietà", "Valore"])
        for k, v in props.items():
            ws.append([k, v])
        wb.save(str(dest))

    def import_from_excel(self, document_id: int, src: Path) -> int:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl non installato")

        wb = openpyxl.load_workbook(str(src))
        ws = wb.active
        props = {}
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if not header_found:
                if row and str(row[0]).lower() == "proprietà":
                    header_found = True
                continue
            if row and row[0]:
                props[str(row[0])] = str(row[1]) if row[1] is not None else ""
        self.save_properties(document_id, props)
        return len(props)
